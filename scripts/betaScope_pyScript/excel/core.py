"""
Excel formating and I/O.
"""

import logging, coloredlogs

logging.basicConfig()
log = logging.getLogger(__name__)
coloredlogs.install(level="CRITICAL", logger=log)

import openpyxl


class ExcelBase(object):
    """
    Excel base class with backend openyxl.
    """

    def __init__(self):
        self.workbook = None
        self.counter = 0

    @classmethod
    def open(cls, excel_file):
        load_wb = cls()
        load_wb.workbook = openpyxl.load_workbook(excel_file)
        return load_wb

    @classmethod
    def new(cls):
        new_wb = cls()
        new_wb.workbook = openpyxl.Workbook()

    def __getitem__(self, key):
        """
        dunder method wrapper for self.workbonk (obj:Workbook)

        Agrs:
            key (str) : name of the excel sheet

        Returns:
            return (worksheet, max row)
        """
        return (self.workbook[key], self.workbook[key].max_row)

    def __setitem__(self, key, value):
        self.workbook[key] = value
