from openpyxl import Workbook, load_workbook
import configparser
import os
import json

par_list = [
"runNumber",
"SensorName",
"Temp",
"Bias",
"Resistance",
"pulseArea",
"pulseArea_Error",
"Pmax",
"Pmax_Error",
"RMS",
"RMS_Error",
"Rise_Time",
"Rise_Time_Error",
"dvdt",
"dvdt_Error",
"FWHM",
"FWHM_Error",
"NewPulseArea",
"NewPulseArea_Error",
"FallTime",
"FallTime_Error"
]

par_dict = {
"runNumber" : "A",
"SensorName" : "B",
"Temp" : "G",
"Bias" : "H",
"Resistance" : "L",
"pulseArea" : "J",
"pulseArea_Error" : "K",
"Pmax" : "R",
"Pmax_Error" : "S",
"RMS" : "T",
"RMS_Error" : "U",
"Rise_Time" : "Z",
"Rise_Time_Error" : "AA",
"dvdt" : "AB",
"dvdt_Error" : "AC",
"FWHM" : "AL",
"FWHM_Error" : "AM",
"NewPulseArea" : "CA",
"NewPulseArea_Error" : "CB",
"FallTime" : "DG",
"FallTime_Error" : "DH",
"cycle" : "F",
"CFD50Time" : "BW",
"CFD50Time_Err" : "BX",
"CFD20Time" : "DD",
"CFD20Time_Err" : "DE",
"Leakage" : "C"
}

def mergeExcel(fname="_results.xlxs"):
    output_dir = os.environ['BETASCOPE_SCRIPTS']
    if not output_dir:
        output_dir = "/tmp/"

    if os.path.exists("{}/user_data/".format(output_dir)):
        pass
    else:
        os.makedirs("{}/user_data/".format(output_dir))

    no_merge_file = False
    if os.path.exists("{}/user_data/merged_beta_results.xlsx".format(output_dir)):
        src_wb = load_workbook("{}/user_data/merged_beta_results.xlsx".format(output_dir))
    else:
        no_merge_file = True
        src_wb = Workbook()
        src_wb.create_sheet("DUT")
        src_wb.create_sheet("TRIG")

    input_wb = load_workbook(fname)

    # for meta data
    input_run_number = input_wb["DUT"][par_dict["runNumber"]+str(1)].value.split("->")[0]
    input_sensor_name = input_wb["DUT"][par_dict["SensorName"]+str(1)].value
    start_row_in_merge = None
    end_row_in_merge = None

    sheets = ["DUT", "TRIG"]
    for sheet in sheets:
        src_ws = src_wb[sheet]
        input_ws = input_wb[sheet]
        input_max_row = input_ws.max_row
        max_row = src_ws.max_row
        for par in par_list:
            rowCounter = 2
            for row in range(1,input_max_row+1):
                if start_row_in_merge==None:
                    start_row_in_merge = max_row+rowCounter
                input_cell = par_dict[par] + str(row)
                src_cell = par_dict[par] + str(max_row+rowCounter)
                src_ws[src_cell] = input_ws[input_cell].value
                end_row_in_merge = max_row+rowCounter
                rowCounter+=1

    src_wb.save("{}/user_data/merged_beta_results.xlsx".format(output_dir))

    if os.path.exists("{}/user_data/merged_log.json".format(output_dir)):
        with open("{}/user_data/merged_log.json".format(output_dir)) as f:
            print(f)
            meta_data = json.load(f)
            keyName = "run"+str(input_run_number)
            duplicate = 2
            while keyName in meta_data.keys():
                keyName = keyName.split("p")[0]+"p{}".format(duplicate)
                duplicate+=1
            meta_data[keyName] = {"sensor":str(input_sensor_name), "start":start_row_in_merge, "end":end_row_in_merge}
        with open("{}/user_data/merged_log.json".format(output_dir),"w") as newf:
            json.dump(meta_data,newf)
    else:
        with open("{}/user_data/merged_log.json".format(output_dir), "w") as f:
            meta_data = {}
            meta_data["run"+str(input_run_number)] = {"sensor":str(input_sensor_name), "start":start_row_in_merge, "end":end_row_in_merge}
            json.dump(meta_data, f)


def parseINIToExcel(fname="_results.ini"):
    config = configparser.ConfigParser()
    config.read(fname)
    config_section = config.sections()
    print(config_section)

    description_file = None
    for descr in os.listdir("./"):
        if "_Description.ini" in descr:
            description_file = configparser.ConfigParser()
            description_file.read(descr)
            print("found DAQ description file")
            break

    wb = Workbook()
    wb.create_sheet("DUT")
    wb.create_sheet("TRIG")
    #ws = wb.active

    rowCounter = 1
    dut_trig = ["DUT", "Trig"]

    RunNum = 100
    SensorName = "Hi"
    Temp = 20
    trigBias = 395

    if description_file:
        try:
            RunNum = description_file["Run_Description"]["Run_Number"]

            SensorName = description_file["Run_Description"]["DUT_Senor_Name"]
            SensorName += "-Fluence "+description_file["Run_Description"]["DUT_Fluence_Type"]+"-"+description_file["Run_Description"]["DUT_Fluence"]
            SensorName += "--"+description_file["Run_Description"]["DUT_Readout_Board"]+"-"+description_file["Run_Description"]["DUT_Readout_Board_Number"]

            Temp = description_file["Run_Description"]["Temperature"]
            trigBias = description_file["Run_Description"]["Trigger_Voltage"]
        except Exception as e:
            print(Exception)
            RunNum = 100
            SensorName = "Hi"
            Temp = 20
            trigBias = 395

    Resistance = 4700

    for ch in dut_trig:
        rowCounter = 1
        for bias in config_section:
            myRunNum = str(RunNum)+"->"+str(rowCounter)
            if ch in bias:
                if ch != "Trig":
                    ws = wb["DUT"]
                    if ".." in bias:
                        Bias = bias[bias.find("_")+1:bias.find("V")]
                        cycle = bias.split("..")[1]
                        if "_" in cycle:
                            cycle = cycle.split("_")[0]
                    else:
                        Bias = bias[bias.find("_")+1:bias.find("V")]
                        cycle = 1
                else:
                    #Bias = trigBias
                    ws = wb["TRIG"]
                    try:
                        SensorName = description_file["Run_Description"]["Trigger_Sensor_Name"]
                    except:
                        pass
                    try:
                        Bias = config[bias]["trigger_bias"]
                        if ".." in bias:
                            cycle = bias.split("..")[1]
                            if "_" in cycle:
                                cycle = cycle.split("_")[0]
                        else:
                            cycle = 1
                    except:
                        Bias = -390
                        cycle = 1
                for par in par_list:
                    if (par == "SensorName"):
                        cell = par_dict[par] + str(rowCounter)
                        ws[cell] = SensorName
                    elif(par=="runNumber"):
                        cell = par_dict[par] + str(rowCounter)
                        ws[cell] = myRunNum
                    elif (par == "Temp"):
                        try:
                            Temp = config[bias]["temperature"]
                        except:
                            Temp = "-30"
                        cell = par_dict[par] + str(rowCounter)
                        ws[cell] = float(Temp)
                    elif (par == "Bias"):
                        cell = par_dict[par] + str(rowCounter)
                        ws[cell] = float(Bias)
                    elif par=="cycle":
                            cell = par_dict["cycle"] + str(rowCounter)
                            ws[cell] = int(cycle)
                    elif (par == "Resistance"):
                        cell = par_dict[par] + str(rowCounter)
                        ws[cell] = float(Resistance)
                    else:
                        cell = par_dict[par] + str(rowCounter)
                        ws[cell] = float(config[bias][par])
                rowCounter+=1
        rowCounter+=1

    wb.save("_results.xlsx")

    mergeExcel("_results.xlsx")

#===============================================================================
def injectData( paramName ):
    import subprocess

    meta_data = None
    src_wb = None

    output_dir = os.environ["BETASCOPE_SCRIPTS"]

    if os.path.exists("{}/user_data/merged_log.json".format(output_dir)):
        with open("{}/user_data/merged_log.json".format(output_dir)) as f:
            meta_data = json.load(f)
            print(meta_data)
        src_wb = load_workbook("{}/user_data/merged_beta_results.xlsx".format(os.environ["BETASCOPE_SCRIPTS"]) )
    else:
        print("Cannot find merge excel file location")
        return -1

    if paramName=="timing":
        p = subprocess.call("python2 $BETASCOPE_SCRIPTS/betaScope_pyScript/get_time_res.py --CFD 50", shell=True)
        p = subprocess.call("python2 $BETASCOPE_SCRIPTS/betaScope_pyScript/get_time_res.py --CFD 20", shell=True)

        with open("res50.txt") as f:
            start_row = None
            end_row = None
            for line in f.readlines():
                raw_txt_data = line.split(",")
                if start_row == None:
                    for key in meta_data.keys():
                        if raw_txt_data[0] in key:
                            start_row = meta_data["run"+str(raw_txt_data[0])]["start"]
                            break
                src_wb["DUT"][par_dict["CFD50Time"]+str(start_row)] = float(raw_txt_data[3]) # stroing timing res
                src_wb["DUT"][par_dict["CFD50Time_Err"]+str(start_row)] = float(raw_txt_data[4])
                start_row+=1

        with open("res20.txt") as f:
            start_row = None
            end_row = None
            for line in f.readlines():
                raw_txt_data = line.split(",")
                if start_row == None:
                    for key in meta_data.keys():
                        if raw_txt_data[0] in key:
                            start_row = meta_data["run"+str(raw_txt_data[0])]["start"]
                            break
                src_wb["DUT"][par_dict["CFD20Time"]+str(start_row)] = float(raw_txt_data[3]) # stroing timing res
                src_wb["DUT"][par_dict["CFD20Time_Err"]+str(start_row)] = float(raw_txt_data[4])
                start_row+=1
        src_wb.save("{}/user_data/merged_beta_results.xlsx".format(os.environ["BETASCOPE_SCRIPTS"]) )

    if paramName=="leakage":
        p = subprocess.call("python2 $BETASCOPE_SCRIPTS/betaScope_pyScript/read_current.py", shell=True)
        with open("leakage.txt") as f:
            start_row = None
            end_row = None
            for line in f.readlines():
                raw_txt_data = line.split(",")
                if start_row == None:
                    print(meta_data.keys())
                    print(raw_txt_data[0])
                    for key in meta_data.keys():
                        if str(raw_txt_data[0]) in key:
                            start_row = meta_data["run"+str(raw_txt_data[0])]["start"]
                            print("starting row {}".format(start_row))
                            break
                    if start_row==None:
                        print("cannot find starting row")
                        return -1
                src_wb["DUT"][par_dict["Leakage"]+str(start_row)] = float(raw_txt_data[3]) # stroing timing res
                #print(raw_txt_data[3])
                start_row+=1
        src_wb.save("{}/user_data/merged_beta_results.xlsx".format(os.environ["BETASCOPE_SCRIPTS"]) )


#===============================================================================
def toROOT():
    import ROOT
    from array import array

    src_wb = None
    output_dir = os.environ["BETASCOPE_SCRIPTS"]
    if os.path.exists("{}/user_data/merged_log.json".format(output_dir)):
        with open("{}/user_data/merged_log.json".format(output_dir)) as f:
            meta_data = json.load(f)
        src_wb = load_workbook("{}/user_data/merged_beta_results.xlsx".format(os.environ["BETASCOPE_SCRIPTS"]) )
    else:
        print("Cannot find merge excel file location")
        return -1

    src_ws = src_wb["DUT"]
    tfile = ROOT.TFile("{}/user_data/merged.root".format(output_dir), "UPDATE")
    tfile.cd()
    for key in meta_data.keys():
        start_row = meta_data[key]["start"]
        end_row = meta_data[key]["end"]
        ttree = ROOT.TTree(key,"from merged excel")
        branches = {}
        for par in par_dict.keys():
            if "SensorName" in par:
                branches[par] = array("c", str(src_ws[par_dict[par]+str(start_row)].value))
                ttree.Branch(par, branches[par], "{}/C".format(par) )
            elif "runNumber" in par:
                continue
            else:
                branches[par] = array("d", [0])
                ttree.Branch(par, branches[par], "{}/D".format(par) )
        for row in range(start_row,end_row+1):
            for par in par_dict.keys():
                if "SensorName" in par or "runNumber" in par:
                    continue
                else:
                    if src_ws[par_dict[par]+str(row)].value == None:
                        src_ws[par_dict[par]+str(row)].value = -99999999999999
                    branches[par][0] = float(src_ws[par_dict[par]+str(row)].value)
            ttree.Fill()
        ttree.Write(key, ROOT.TObject.kOverwrite)
    tfile.Close()


#===============================================================================
if __name__=="__main__":
    import argparse
    cml_parser = argparse.ArgumentParser()
    cml_parser.add_argument("-task", dest="task", nargs="?", default="parse", type=str, help="no help!")
    cml_parser.add_argument("-param", dest="param", nargs="?", default="timing", type=str, help="parameters name!")
    argv = cml_parser.parse_args()

    if argv.task == "parse":
        parseINIToExcel()
        injectData("timing")
        injectData("leakage")
    elif argv.task=="merge":
        mergeExcel()
    elif argv.task=="inject":
        injectData(argv.param)
    elif argv.task=="toROOT":
        toROOT()
    else:
        parseINIToExcel()
