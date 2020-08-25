from openpyxl import Workbook, load_workbook
import configparser
import os
import json

from get_time_res import Get_Time_Resolution
from read_current import Read_Current
from daq_info import DAQInfo
from config_reader import RUN_INFO_VER

import sys
sys.path.append(f'{os.environ["BETASCOPE_SCRIPTS"]}/UDI_reader/')
from UDI_reader import UDI_reader

import logging

logging.basicConfig()
# logging.getLogger().setLevel(logging.CRITICAL)
log = logging.getLogger(__name__)

par_list = [
    "runNumber",
    "SensorName",
    "Temp",
    "Bias",
    "Resistance",
    "pin_charge",
    "foot_cv",
    "pulseArea",
    "pulseArea_Error",
    "Pmax",
    "Pmax_Error",
    "RMS",
    "RMS_Error",
    "Rise_Time",
    "Rise_Time_Error",
    "dvdt",
    "dvdt_Error",
    "FWHM",
    "FWHM_Error",
    "NewPulseArea",
    "NewPulseArea_Error",
    "collected_charge",
    "gain",
    "gain2",
    "FallTime",
    "FallTime_Error",
    "cycle",
]

# name_ref : (ini_key, excel_col)
# if ini_key or excel_col is None, the corresponding data in ini or excel
# will not be parsed
# if the ini_key dose not exist, it will be skipped
INI_TO_EXCEL = {
    "run_number": ("runNumber", "A"),
    "sensor_name": ("SensorName", "B"),
    "temperature": ("Temp", "G"),
    "bias_voltage": ("Bias", "H"),
    "resistance": ("Resistance", "L"),
    "pin_charge": ("pin_charge", "P"),
    "foot_cv": ("foot_cv", "AT"),
    "pulse_area": ("pulseArea", "J"),
    "pulse_area_chi": ("pulseArea_CHI_NDF", None),
    "pulse_area_error": ("pulseArea_Error", "K"),
    "pmax": ("Pmax", "R"),
    "pmax_chi": ("Pmax_CHI_NDF", None),
    "pmax_error": ("Pmax_Error", "S"),
    "rms": ("RMS", "T"),
    "rms_chi": ("RMS_CHI_NDF", None),
    "rms_error": ("RMS_Error", "U"),
    "rise_time": ("Rise_Time", "Z"),
    "rise_time_chi": ("Rise_Time_CHI_NDF", None),
    "rise_time_error": ("Rise_Time_Error", "AA"),
    "dvdt": ("dvdt", "AB"),
    "dvdt_chi": ("dvdt_CHI_NDF", None),
    "dvdt_Error": ("dvdt_Error", "AC"),
    "fwhm": ("FWHM", "AL"),
    "fwhm_chi": ("FWHM_CHI", None),
    "fwhm_error": ("FWHM_Error", "AM"),
    "new_pulse_area": ("NewPulseArea", "CA"),
    "new_pulse_area_chi": ("NewPulseArea_CHI_NDF", None),
    "new_pulse_area_error": ("NewPulseArea_Error", "CB"),
    "collected_charge": ("collected_charge", "CC"),
    "gain": ("gain", "Q"),
    "gain2": ("gain2", "CF"),
    "fall_time": ("FallTime", "DG"),
    "fall_time_chi": ("FallTime_CHI_NDF", None),
    "fall_time_error": ("FallTime_Error", "DH"),
    "cycle": ("cycle", "F"),
    "time_resolution_50": (None, "BW"),
    "time_resolution_50_err": (None, "BX"),
    "time_resolution_20": (None, "DD"),
    "time_resolution_20_err": (None, "DE"),
    "time_resolution_tmax": (None, None),
    "time_resolution_tmax_err": (None, None),
    "time_resolution_fit_tmax": (None, None),
    "time_resolution_fit_tmax_err": (None, None),
    "time_resolution_fit_zero_x_tmax": (None, None),
    "time_resolution_fit_zero_x_tmax_err": (None, None),
    "leakage": ("Leakage", "C"),
}


def MergeExcel(ifile="_results.xlxs"):
    """
    function for read in existing excel file and merge data from ifile
    """

    # loading envir path
    output_dir = os.environ["BETASCOPE_SCRIPTS"]
    if not output_dir:
        output_dir = "/tmp/"

    if not os.path.exists(f"{output_dir}/user_data/"):
        os.makedirs(f"{output_dir}/user_data/")

    # checking merged beta result
    if os.path.exists(f"{output_dir}/user_data/merged_beta_results.xlsx"):
        no_merge_file = False
        merged_wb = load_workbook(f"{output_dir}/user_data/merged_beta_results.xlsx")
    else:
        no_merge_file = True
        merged_wb = Workbook()
        merged_wb.create_sheet("DUT")
        merged_wb.create_sheet("TRIG")

    input_wb = load_workbook(ifile)

    # for meta data
    input_dut_wp = input_wb["DUT"]
    run_number_cell = f"{INI_TO_EXCEL['run_number'][1]}1"
    sensor_name_cell = f"{INI_TO_EXCEL['sensor_name'][1]}1"
    input_run_number = input_dut_wp[run_number_cell].value.split("->")[0]
    input_sensor_name = input_dut_wp[sensor_name_cell].value

    input_max_row = input_wb["DUT"].max_row
    merged_max_row = merged_wb["DUT"].max_row

    # check meta data
    if os.path.exists(f"{output_dir}/user_data/merged_log.json"):
        with open(f"{output_dir}/user_data/merged_log.json") as f:
            meta_data = json.load(f)
            keyName = f"run{input_run_number}"
            if keyName in meta_data:
                start_row_in_merge = meta_data[keyName]["start"]
                end_row_in_merge = meta_data[keyName]["end"]
                diff = end_row_in_merge - start_row_in_merge
                if input_max_row > diff:
                    new_rows = input_max_row - diff
                else:
                    new_rows = 0
                end_row_in_merge += new_rows
            else:
                new_rows = 0
                start_row_in_merge = merged_max_row + 2
                end_row_in_merge = start_row_in_merge + input_max_row
            meta_data[f"run{input_run_number}"] = {
                "sensor": f"{input_sensor_name}",
                "start": start_row_in_merge,
                "end": end_row_in_merge,
            }
        with open(f"{output_dir}/user_data/merged_log.json", "w") as newf:
            json.dump(meta_data, newf)
    else:
        new_rows = 0
        with open(f"{output_dir}/user_data/merged_log.json", "w") as f:
            start_row_in_merge = merged_max_row + 1
            end_row_in_merge = start_row_in_merge + input_max_row
            meta_data = {}
            meta_data[f"run{input_run_number}"] = {
                "sensor": f"{input_sensor_name}",
                "start": start_row_in_merge,
                "end": end_row_in_merge,
            }
            json.dump(meta_data, f)

    sheets = ["DUT", "TRIG"]
    for sheet in sheets:
        merged_ws = merged_wb[sheet]
        input_ws = input_wb[sheet]
        for i in range(new_rows):
            merged_ws.insert(end_row_in_merge)
        for par in INI_TO_EXCEL.keys():
            if INI_TO_EXCEL[par][1] == None:
                continue
            rowCounter = start_row_in_merge
            for row in range(1, input_max_row + 1):
                input_cell = f"{INI_TO_EXCEL[par][1]}{row}"
                merged_cell = f"{INI_TO_EXCEL[par][1]}{rowCounter}"
                merged_ws[merged_cell] = input_ws[input_cell].value
                rowCounter += 1

    merged_wb.save(f"{output_dir}/user_data/merged_beta_results.xlsx")


def ParseINIToExcel(fname="_results.ini", update_merge=True):
    """
    Parse data in ini file to excel
    """
    data_ini = configparser.ConfigParser()
    data_ini.read(fname)
    data_ini_section = data_ini.sections()

    # creating new work book and pages
    wb = Workbook()
    wb.create_sheet("DUT")
    wb.create_sheet("TRIG")
    # ws = wb.active

    # look for description file generated by the DAQ
    description_file = DAQInfo()
    for descr in os.listdir("./"):
        if "_Description.ini" in descr:
            description_file = DAQInfo.open(descr)
            log.info("found DAQ description file")
            break

    sensor_name = description_file.full_name

    # total transipedence (include amp)
    resistance = 4700.

    # Get stuff from UDI file
    try:
        UDI_number = description_file.dut_udi
        reader = UDI_reader()
        pin_charge = reader.get_pin_charge(UDI_number)
        foot_cv = reader.get_foot(UDI_number)
    except:
        pin_charge = 0.5
        foot_cv = 0.

    # start writing data to excel workbook
    for ch in ["DUT", "Trig"]:
        row = 1
        for bias in data_ini_section:

            ini_key = INI_TO_EXCEL["new_pulse_area"][0]
            value = data_ini[bias][ini_key]
            collected_charge = float(value)/resistance
            gain = collected_charge/pin_charge

            my_run_num = f"{description_file.run_number}->{row}"
            if ch in bias:
                if ch == "DUT":
                    ws = wb["DUT"]
                    run_header = bias.split(",")
                    my_bias = run_header[1]
                    cycle = run_header[2]
                    my_sensor_name = sensor_name
                else:
                    ws = wb["TRIG"]
                    my_sensor_name = description_file.trig_name
                    run_header = bias.split(",")
                    my_bias = run_header[1].replace("V", "")
                    cycle = run_header[2]
                for par in INI_TO_EXCEL.keys():
                    cell = f"{INI_TO_EXCEL[par][1]}{row}"
                    if par == "sensor_name":
                        ws[cell] = my_sensor_name
                    elif par == "run_number":
                        ws[cell] = my_run_num
                    elif par == "temperature":
                        ws[cell] = description_file.temperature
                    elif par == "bias_voltage":
                        try:
                            ws[cell] = float(my_bias)
                        except:
                            if "V" in my_bias:
                                my_bias = my_bias.replace("V", "")
                                ws[cell] = float(my_bias)
                            else:
                                ws[cell] = my_bias
                    elif par == "cycle":
                        ws[cell] = int(cycle)
                    elif par == "resistance":
                        ws[cell] = float(resistance)
                    elif par == "pin_charge":
                        ws[cell] = float(pin_charge)
                    elif par == "collected_charge":
                        ws[cell] = float(collected_charge)
                    elif par == "gain":
                        ws[cell] = float(gain)
                    elif par == "gain2":
                        ws[cell] = float(gain)
                    elif par == "foot_cv":
                        ws[cell] = float(foot_cv)
                    else:
                        if None in INI_TO_EXCEL[par]:
                            continue
                        try:
                            ini_key = INI_TO_EXCEL[par][0]
                            value = data_ini[bias][ini_key]
                            ws[cell] = float(value)
                        except:
                            continue
                row += 1
        row += 1  # skip one line

    end_row = len(data_ini_section)

    log.info("Getting time resolution")

    my_trig_name = description_file.trig_name.lower()
    if "hpk" in my_trig_name and "s8664" in my_trig_name:
        my_trig_name = "hpks8664"

    res50_result = Get_Time_Resolution(
        f"run_info_v{RUN_INFO_VER}.ini",
        "50",
        description_file.scope.lower(),
        my_trig_name,
        description_file.run_number,
    )
    res = {}
    res["time_resolution_50"] = {key: item[3] for key, item in res50_result.items()}
    res["time_resolution_50_err"] = {key: item[4] for key, item in res50_result.items()}
    # res["cycle"] = [item[5] for item in res50_result]
    # res["Bias"] = [item[2] for item in res50_result]
    InjectSortColData(
        wb["DUT"], 1, "time_resolution_50", ["bias_voltage", "cycle"], res
    )
    InjectSortColData(
        wb["DUT"], 1, "time_resolution_50_err", ["bias_voltage", "cycle"], res
    )

    res20_result = Get_Time_Resolution(
        f"run_info_v{RUN_INFO_VER}.ini",
        "20",
        description_file.scope.lower(),
        my_trig_name,
        description_file.run_number,
    )
    res = {}
    res["time_resolution_20"] = {key: item[3] for key, item in res20_result.items()}
    res["time_resolution_20_err"] = {key: item[4] for key, item in res20_result.items()}
    InjectSortColData(
        wb["DUT"], 1, "time_resolution_20", ["bias_voltage", "cycle"], res
    )
    InjectSortColData(
        wb["DUT"], 1, "time_resolution_20_err", ["bias_voltage", "cycle"], res
    )

    leakage_data = Read_Current(f"run_info_v{RUN_INFO_VER}.ini")
    leakage_current = {}
    leakage_current["leakage"] = {
        key: item[3] * 1000 for key, item in leakage_data.items()
    }
    InjectSortColData(
        wb["DUT"], 1, "leakage", ["bias_voltage", "cycle"], leakage_current
    )

    wb.save("_results.xlsx")

    if update_merge:
        MergeExcel("_results.xlsx")


# ===============================================================================
def InjectColData(ws, start_row, par, data_list):
    """
    Method for injecting data
    """
    row = start_row
    for data in data_list:
        cell = f"{par}{row}"
        ws[cell] = data
        row += 1


def InjectSortColData(ws, start_row, par, sort_pars, data_list):
    """
    Similar to InjectColData. data_list is tuple (data, data for sorting)
    data_list[par] needs to be dict, the key is tuple with values from sort_pars
    """
    if INI_TO_EXCEL[par][1] == None:
        raise ValueError(f"excel col cannot be None for {par}")

    sorting_buffer = []
    row = start_row
    for i in range(len(data_list[par])):
        buffer = []
        for sort_par in sort_pars:
            if INI_TO_EXCEL[sort_par][1] == None:
                raise ValueError(f"excel col cannot be None for {par}")
            cell = f"{INI_TO_EXCEL[sort_par][1]}{row}"
            buffer.append(ws[cell].value)
        sorting_buffer.append((row, tuple(buffer)))
        row += 1

    if isinstance(data_list[par], dict):
        for row, key in sorting_buffer:
            ws[f"{INI_TO_EXCEL[par][1]}{row}"] = data_list[par][key]
    else:
        pass


# ===============================================================================
def toROOT():
    import ROOT
    from array import array

    src_wb = None
    output_dir = os.environ["BETASCOPE_SCRIPTS"]
    if os.path.exists("{}/user_data/merged_log.json".format(output_dir)):
        with open("{}/user_data/merged_log.json".format(output_dir)) as f:
            meta_data = json.load(f)
        src_wb = load_workbook(
            "{}/user_data/merged_beta_results.xlsx".format(
                os.environ["BETASCOPE_SCRIPTS"]
            )
        )
    else:
        print("Cannot find merge excel file location")
        return -1

    src_ws = src_wb["DUT"]
    tfile = ROOT.TFile("{}/user_data/merged.root".format(output_dir), "UPDATE")
    tfile.cd()
    for key in meta_data.keys():
        start_row = meta_data[key]["start"]
        end_row = meta_data[key]["end"]
        ttree = ROOT.TTree(key, "from merged excel")
        branches = {}
        for par in par_dict.keys():
            if "SensorName" in par:
                branches[par] = array(
                    "c", str(src_ws[par_dict[par] + str(start_row)].value)
                )
                ttree.Branch(par, branches[par], "{}/C".format(par))
            elif "runNumber" in par:
                continue
            else:
                branches[par] = array("d", [0])
                ttree.Branch(par, branches[par], "{}/D".format(par))
        for row in range(start_row, end_row + 1):
            for par in par_dict.keys():
                if "SensorName" in par or "runNumber" in par:
                    continue
                else:
                    if src_ws[par_dict[par] + str(row)].value == None:
                        src_ws[par_dict[par] + str(row)].value = -99999999999999
                    branches[par][0] = float(src_ws[par_dict[par] + str(row)].value)
            ttree.Fill()
        ttree.Write(key, ROOT.TObject.kOverwrite)
    tfile.Close()


# ===============================================================================
if __name__ == "__main__":
    import argparse

    cml_parser = argparse.ArgumentParser()
    cml_parser.add_argument(
        "-task", dest="task", nargs="?", default="parse", type=str, help="no help!"
    )
    cml_parser.add_argument(
        "-param",
        dest="param",
        nargs="?",
        default="timing",
        type=str,
        help="parameters name!",
    )
    argv = cml_parser.parse_args()

    if argv.task == "parse":
        ParseINIToExcel()
    elif argv.task == "merge":
        MergeExcel()
    elif argv.task == "toROOT":
        toROOT()
    else:
        ParseINIToExcel()
